from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

class ExportManager:
    def __init__(self):
        # Use default Helvetica font (supports basic characters)
        self.font_name = 'Helvetica'
            
    def export_pdf(self, quotation, file_path):
        doc = SimpleDocTemplate(file_path, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, 
                               topMargin=1.5*cm, bottomMargin=1.5*cm)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        company_style = ParagraphStyle(
            'Company',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=5,
            alignment=0,
            fontName=self.font_name
        )
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            alignment=1,
            fontName=self.font_name
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=9,
            spaceAfter=2,
            fontName=self.font_name
        )
        
        label_style = ParagraphStyle(
            'Label',
            parent=styles['Normal'],
            fontSize=10,
            fontName=self.font_name,
            alignment=0
        )
        
        value_style = ParagraphStyle(
            'Value',
            parent=styles['Normal'],
            fontSize=10,
            fontName=self.font_name,
            alignment=0
        )
        
        # Header section - Company info on left, Title on right
        header_layout = [
            [
                Paragraph("MEFE MAKİNA", company_style),
                Paragraph("", ParagraphStyle('Empty', fontSize=1)),
                Paragraph("TEKLİF", title_style)
            ],
            [
                Paragraph("Organize Sanayi Bölgesi", header_style),
                Paragraph("", ParagraphStyle('Empty', fontSize=1)),
                Paragraph("", ParagraphStyle('Empty', fontSize=1))
            ],
            [
                Paragraph("Tel: +90 XXX XXX XX XX", header_style),
                Paragraph("", ParagraphStyle('Empty', fontSize=1)),
                Paragraph("", ParagraphStyle('Empty', fontSize=1))
            ],
            [
                Paragraph("E-mail: info@mefemakina.com", header_style),
                Paragraph("", ParagraphStyle('Empty', fontSize=1)),
                Paragraph("", ParagraphStyle('Empty', fontSize=1))
            ]
        ]
        
        header_table = Table(header_layout, colWidths=[7*cm, 1*cm, 7*cm])
        header_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Quotation info section
        company_name = quotation.get('company_name', '')
        info_data = [
            [Paragraph("Teklif No:", label_style), Paragraph(quotation['quotation_no'], value_style)],
            [Paragraph("Tarih:", label_style), Paragraph(quotation['date'], value_style)],
            [Paragraph("Firma:", label_style), Paragraph(company_name, value_style)],
            [Paragraph("Teslimat:", label_style), Paragraph(quotation['delivery_type'], value_style)]
        ]
        
        info_table = Table(info_data, colWidths=[3*cm, 12*cm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Items table
        if quotation['items']:
            items_data = [['Sıra', 'Kod', 'Açıklama', 'Miktar', 'Birim', 'Birim Fiyat', 'Tutar']]
            
            for idx, item in enumerate(quotation['items'], 1):
                items_data.append([
                    str(idx),
                    item['code'],
                    item['description'],
                    str(item['quantity']),
                    item['unit'],
                    f"{float(item['unit_price']):.2f}",
                    f"{float(item['total']):.2f}"
                ])
            
            items_table = Table(items_data, colWidths=[0.8*cm, 2.5*cm, 6*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.5*cm])
            items_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), self.font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.whitesmoke]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            story.append(items_table)
            story.append(Spacer(1, 0.5*cm))
        
        # Total row
        total_data = [
            ['', '', '', '', '', Paragraph('Toplam:', label_style), Paragraph(f"{float(quotation['total_amount']):.2f} {quotation['currency']}", value_style)]
        ]
        total_table = Table(total_data, colWidths=[0.8*cm, 2.5*cm, 6*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2.5*cm])
        total_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (5, 0), (5, 0), 'RIGHT'),
            ('ALIGN', (6, 0), (6, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(total_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Notes
        if quotation['notes']:
            story.append(Paragraph("Notlar:", label_style))
            story.append(Paragraph(quotation['notes'], header_style))
            story.append(Spacer(1, 0.5*cm))
        
        # Signature section
        signature_data = [
            [Paragraph("Kabul Eden:", label_style), Paragraph("", ParagraphStyle('Empty', fontSize=1)), Paragraph("Teklif Veren:", label_style)],
            [Paragraph("", ParagraphStyle('Empty', fontSize=1)), Paragraph("", ParagraphStyle('Empty', fontSize=1)), Paragraph("", ParagraphStyle('Empty', fontSize=1))],
            [Paragraph("", ParagraphStyle('Empty', fontSize=1)), Paragraph("", ParagraphStyle('Empty', fontSize=1)), Paragraph("", ParagraphStyle('Empty', fontSize=1))],
            [Paragraph("İmza:", label_style), Paragraph("", ParagraphStyle('Empty', fontSize=1)), Paragraph("İmza:", label_style)]
        ]
        
        signature_table = Table(signature_data, colWidths=[7*cm, 1*cm, 7*cm])
        signature_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWHEIGHT', (1, 0), (1, -1), 2*cm),
        ]))
        story.append(signature_table)
        
        # Build PDF
        doc.build(story)
        
    def export_excel(self, quotation, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teklif"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        # Header information
        ws['A1'] = "TEKLİF"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')
        
        ws['A3'] = "Teklif No:"
        ws['B3'] = quotation['quotation_no']
        ws['A4'] = "Tarih:"
        ws['B4'] = quotation['date']
        ws['A5'] = "Teslimat Tipi:"
        ws['B5'] = quotation['delivery_type']
        ws['A6'] = "Para Birimi:"
        ws['B6'] = quotation['currency']
        
        # Items table header
        headers = ['Sıra', 'Kod', 'Açıklama', 'Miktar', 'Birim', 'Birim Fiyat', 'Toplam']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_align
            
        # Items data
        for idx, item in enumerate(quotation['items'], 1):
            row = 8 + idx
            ws.cell(row=row, column=1, value=idx).alignment = center_align
            ws.cell(row=row, column=2, value=item['code']).border = border
            ws.cell(row=row, column=3, value=item['description']).border = border
            ws.cell(row=row, column=4, value=item['quantity']).border = border
            ws.cell(row=row, column=4).alignment = center_align
            ws.cell(row=row, column=5, value=item['unit']).border = border
            ws.cell(row=row, column=5).alignment = center_align
            ws.cell(row=row, column=6, value=item['unit_price']).border = border
            ws.cell(row=row, column=6).alignment = right_align
            ws.cell(row=row, column=7, value=item['total']).border = border
            ws.cell(row=row, column=7).alignment = right_align
            
        # Total row
        total_row = 8 + len(quotation['items']) + 1
        ws.cell(row=total_row, column=6, value="Toplam:").font = header_font
        ws.cell(row=total_row, column=6).alignment = right_align
        ws.cell(row=total_row, column=7, value=quotation['total_amount']).font = header_font
        ws.cell(row=total_row, column=7).alignment = right_align
        
        # Notes
        if quotation['notes']:
            notes_row = total_row + 2
            ws.cell(row=notes_row, column=1, value="Notlar:").font = header_font
            ws.cell(row=notes_row + 1, column=1, value=quotation['notes'])
            ws.merge_cells(f'A{notes_row + 1}:G{notes_row + 1}')
            
        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        # Save
        wb.save(file_path)
